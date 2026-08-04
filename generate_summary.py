import logging
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from scripts.spreadsheet_safety import write_excel_safely

log = logging.getLogger(__name__)

# Set OUTPUT_DIR to the project root's output directory
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")
SUMMARY_FILE = os.path.join(OUTPUT_DIR, "Summary_Report.xlsx")


def get_processed_files(output_dir):
    """Finds and returns all processed Excel files in the output directory."""
    return sorted([f for f in os.listdir(output_dir) if f.endswith("_Processed.xlsx")])


def process_summary_data(output_dir, processed_files):
    """Reads processed files and compiles summary statistics."""
    summary_data = []
    for file in processed_files:
        file_path = os.path.join(output_dir, file)
        try:
            df = pd.read_excel(file_path)
            mean_value = df["Processed_Value"].mean()
            median_value = df["Processed_Value"].median()
            outlier_count = df["Is_Outlier"].sum()
            summary_data.append(
                {
                    "File": file,
                    "Mean_Processed_Value": mean_value,
                    "Median_Processed_Value": median_value,
                    "Outlier_Count": outlier_count,
                }
            )
        except Exception:
            log.exception(f"Internal error processing {file}")
            print(f"Error processing {file}: An unexpected error occurred.")
    return summary_data


def format_summary_excel(summary_file):
    """Formats the summary Excel file by bolding headers, adjusting widths, and adding a chart."""
    wb = load_workbook(summary_file)
    ws = wb.active

    # Bold headers and adjust column widths
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 25

    # Create a bar chart for Outlier_Count
    chart = BarChart()
    chart.title = "Outlier Count per File"
    chart.x_axis.title = "File"
    chart.y_axis.title = "Outlier Count"

    data = Reference(
        ws, min_col=4, min_row=1, max_row=ws.max_row, max_col=4
    )  # Outlier_Count
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # File names
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Place the chart below the data
    ws.add_chart(chart, f"A{ws.max_row + 3}")

    wb.save(summary_file)


def main():
    processed_files = get_processed_files(OUTPUT_DIR)
    if not processed_files:
        print(f"No processed files found in {OUTPUT_DIR}")
        return

    summary_data = process_summary_data(OUTPUT_DIR, processed_files)

    summary_df = pd.DataFrame(summary_data)
    write_excel_safely(summary_df, SUMMARY_FILE, index=False)

    format_summary_excel(SUMMARY_FILE)
    print(f"Summary report with chart saved to: {SUMMARY_FILE}")


if __name__ == "__main__":
    main()
